"""Exporta arquivo de auditoria com os top-200 FPs de score >= 0.90 para revisao humana.

O usuario revisa cada caso e preenche a coluna 'veredito_humano':
  - 'concordo'        -> rotulo INPI estava correto, NAO colidem
  - 'discordo'        -> rotulo INPI estava errado, REALMENTE colidem
  - 'duvidoso'        -> caso ambiguo
  - 'rotulo_invalido' -> dados aparentemente corruptos

Ao final, podemos retroalimentar 'discordo' como label=1 e treinar de novo,
ou usar a taxa de concordancia como teto pratico do modelo.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parent
ART = ROOT / "artifacts"


def main(top_n: int = 200, score_min: float = 0.90) -> None:
    df_full = pd.read_excel(ART / "brand_similarity_input_view_enriched.xlsx")
    print(f"[i] enriched: {len(df_full)} linhas")

    # Localiza coluna do rotulo INPI
    label_col = None
    for c in df_full.columns:
        if c.lower() == "label" or "rotulo" in c.lower():
            label_col = c
            break
    if label_col is None:
        raise RuntimeError("Coluna de rotulo nao encontrada.")
    df_full["y"] = df_full[label_col].astype(int)

    threshold = 0.26
    df_full["pred"] = (df_full["score_nn"] >= threshold).astype(int)
    fp_mask = (df_full["y"] == 0) & (df_full["pred"] == 1) & (df_full["score_nn"] >= score_min)
    fp = df_full[fp_mask].copy().sort_values("score_nn", ascending=False).head(top_n)

    print(f"[i] FPs com score >= {score_min}: {fp_mask.sum()} no total")
    print(f"[i] Exportando os top {len(fp)} para auditoria...")

    cols = [
        "marca_monitorada", "marca_colidente",
        "classe_marca_monitorada", "classe_marca_colidente",
        "especificacao_monitorado", "especificacao_colidente",
        "score_nn", "score_heuristic", "label",
    ]
    cols = [c for c in cols if c in fp.columns]
    out = fp[cols].copy()

    # Renomear para clareza
    rename = {
        "score_nn": "score_modelo",
        "score_heuristic": "score_ofta_heuristico",
        "label": "rotulo_inpi_original",
    }
    out = out.rename(columns=rename)
    out.insert(0, "indice_original", fp.index)

    # Coluna para preenchimento humano
    out["veredito_humano"] = ""        # 'concordo' | 'discordo' | 'duvidoso' | 'rotulo_invalido'
    out["observacao"] = ""

    out_path = ART / "auditoria_top_fps.xlsx"
    out.to_excel(out_path, index=False)

    # Adicionar formatacao basica (largura de coluna + freeze panes via openpyxl)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        ws.freeze_panes = "C2"

        widths = {
            "A": 8, "B": 35, "C": 35,
            "D": 8, "E": 8,
            "F": 60, "G": 60,
            "H": 12, "I": 18, "J": 16,
            "K": 18, "L": 30,
        }
        for col_letter, w in widths.items():
            if col_letter in [c.column_letter for c in ws[1]]:
                ws.column_dimensions[col_letter].width = w

        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

        # Cabecalho destacado
        header_fill = openpyxl.styles.PatternFill(
            "solid", fgColor="305496",
        )
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

        wb.save(out_path)
    except Exception as exc:  # noqa: BLE001
        print(f"[!] Sem formatacao adicional ({exc})")

    print(f"[OK] Arquivo de auditoria salvo: {out_path}")
    print()
    print("=" * 70)
    print("INSTRUCOES PARA O USUARIO")
    print("=" * 70)
    print("Para cada linha do arquivo, preencha a coluna 'veredito_humano' com:")
    print("  concordo        -> rotulo INPI esta certo, marcas NAO colidem")
    print("  discordo        -> rotulo INPI esta errado, marcas REALMENTE colidem")
    print("  duvidoso        -> caso ambiguo / precisa de mais analise")
    print("  rotulo_invalido -> dados corrompidos ou ininteligiveis")
    print()
    print("Ao final, taxa(discordo) >= 15-20% indicara ruido relevante de rotulo")
    print("e justificara enriquecer o conjunto positivo com esses casos.")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--top", type=int, default=200)
    parser.add_argument("--score-min", type=float, default=0.90)
    args = parser.parse_args()
    main(top_n=args.top, score_min=args.score_min)
